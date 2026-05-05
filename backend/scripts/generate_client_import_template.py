from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


HEADERS = [
    "№ пациента",
    "Фамилия",
    "Имя",
    "Отчество",
    "Дата рождения",
    "Пол",
    "Телефон",
    "E-mail",
    "Тип документа",
    "Серия документа",
    "Номер документа",
    "Кем выдан",
    "Дата выдачи",
    "СНИЛС",
    "Регистрация",
    "Адрес проживания",
    "Организация",
    "Категория допуска",
    "№ справки",
    "Примечание",
]

WIDTHS = [14, 18, 18, 20, 16, 12, 18, 24, 18, 18, 18, 28, 16, 18, 28, 28, 28, 22, 16, 36]

EXAMPLE_ROW = [
    "",
    "Иванов",
    "Иван",
    "Иванович",
    "15.04.1987",
    "муж",
    "+7 999 000-00-00",
    "ivanov@example.com",
    "Паспорт РФ",
    "4501",
    "123456",
    "ГУ МВД России",
    "10.10.2020",
    "111-222-333 44",
    "г. Тверь, ул. Лесная, д. 1",
    "г. Тверь, ул. Лесная, д. 1",
    "Завод Пример",
    "Медкомиссия",
    "МК-001",
    "Заполняет клиент",
]


def main() -> None:
    root = Path(__file__).resolve().parents[2]
    output = root / "frontend" / "public" / "demo" / "client-import-template.xlsx"
    output.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Clients"

    for index, (header, width) in enumerate(zip(HEADERS, WIDTHS), start=1):
        cell = worksheet.cell(row=1, column=index, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0C7D7B")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        worksheet.column_dimensions[cell.column_letter].width = width

    for row in worksheet.iter_rows(min_row=2, max_row=300):
        for cell in row:
            cell.number_format = "@"

    validation = DataValidation(type="list", formula1='"муж,жен"', allow_blank=True)
    worksheet.add_data_validation(validation)
    validation.add("F2:F300")
    worksheet.freeze_panes = "A2"

    example_sheet = workbook.create_sheet("Example")
    example_sheet["A1"] = "Как заполнять шаблон"
    example_sheet["A1"].font = Font(bold=True, size=14)
    example_sheet["A3"] = "1. Заполняйте только первый лист Clients."
    example_sheet["A4"] = "2. Один клиент = одна строка."
    example_sheet["A5"] = "3. № пациента можно оставить пустым: система присвоит номер сама."
    example_sheet["A6"] = "4. Дату пишите в формате ДД.ММ.ГГГГ."
    example_sheet["A7"] = "5. Пол: муж или жен."
    example_sheet["A9"] = "Пример строки:"

    for index, header in enumerate(HEADERS, start=1):
        example_sheet.cell(row=11, column=index, value=header).font = Font(bold=True)
        example_sheet.cell(row=12, column=index, value=EXAMPLE_ROW[index - 1])
        example_sheet.column_dimensions[example_sheet.cell(row=11, column=index).column_letter].width = WIDTHS[index - 1]

    workbook.save(output)
    print(output)


if __name__ == "__main__":
    main()
